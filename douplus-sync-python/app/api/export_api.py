"""
订单数据导出API

职责：导出订单列表数据为Excel文件
"""

import logging
from datetime import datetime
from flask import request, Response
from io import BytesIO
from sqlalchemy import text
from app.api import query_bp
from app.api.common import require_auth, error_response
from app.models import SessionLocal

logger = logging.getLogger(__name__)

# 需要安装openpyxl: pip install openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXPORT_AVAILABLE = True
except ImportError:
    EXPORT_AVAILABLE = False
    logger.warning("openpyxl未安装，导出功能不可用")


@query_bp.route('/task/export', methods=['GET'])
@require_auth
def export_task_data():
    """
    导出订单数据为Excel
    
    参数：与task/page接口相同
    - status: 状态筛选
    - accountId: 账号筛选
    - keyword: 视频标题关键词搜索
    - startDate: 开始日期
    - endDate: 结束日期
    
    返回：Excel文件流
    """
    if not EXPORT_AVAILABLE:
        return error_response('导出功能不可用，请联系管理员安装openpyxl'), 500
    
    user_id = request.user_id
    status = request.args.get('status')
    account_id = request.args.get('accountId')
    keyword = request.args.get('keyword')
    start_date = request.args.get('startDate')
    end_date = request.args.get('endDate')
    
    db = SessionLocal()
    try:
        # 构建查询条件
        where_conditions = [
            "o.user_id = :user_id",
            "o.deleted = 0",
            "a.deleted = 0"
        ]
        params = {'user_id': user_id}
        
        if status:
            where_conditions.append("o.status = :status")
            params['status'] = status
        
        if account_id:
            where_conditions.append("o.account_id = :account_id")
            params['account_id'] = int(account_id)
        
        # 优化：支持视频ID精确匹配或标题模糊搜索
        if keyword:
            where_conditions.append("(o.item_id = :keyword OR o.aweme_title LIKE :keyword_like)")
            params['keyword'] = keyword
            params['keyword_like'] = f'%{keyword}%'
        
        if start_date:
            where_conditions.append("DATE(o.order_create_time) >= :start_date")
            params['start_date'] = start_date
        
        if end_date:
            where_conditions.append("DATE(o.order_create_time) <= :end_date")
            params['end_date'] = end_date
        
        where_clause = " AND ".join(where_conditions)
        
        # 查询订单数据（带效果数据）
        data_sql = f"""
            SELECT 
                o.order_id,
                o.item_id,
                o.status,
                o.budget,
                o.aweme_title,
                a.nickname as account_nickname,
                o.order_create_time,
                oa.total_cost,
                oa.total_play,
                oa.total_like,
                oa.total_comment,
                oa.total_share,
                oa.total_follow,
                oa.total_convert,
                oa.play_per_100_cost,
                oa.avg_convert_cost,
                oa.share_rate,
                oa.play_duration_5s
            FROM douplus_order o
            INNER JOIN douyin_account a ON o.account_id = a.id
            LEFT JOIN douplus_order_agg oa ON o.order_id = oa.order_id
            WHERE {where_clause}
            ORDER BY o.order_create_time DESC
            LIMIT 10000
        """
        
        results = db.execute(text(data_sql), params).fetchall()
        
        # 创建Excel工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "订单数据"
        
        # 设置标题行样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 定义表头
        headers = [
            "订单ID", "视频ID", "状态", "预算(元)", "视频标题", "账号",
            "下单时间", "实际消耗(元)", "播放量", "点赞数", "评论数",
            "转发数", "关注数", "转化数", "百播放量", "转化成本(元)",
            "百转发率", "5秒完播率"
        ]
        
        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # 状态映射
        status_map = {
            'UNPAID': '未支付',
            'AUDITING': '审核中',
            'DELIVERING': '投放中',
            'DELIVERIED': '已完成',
            'UNDELIVERIED': '投放终止',
            'AUDIT_PAUSE': '审核暂停',
            'AUDIT_REJECTED': '审核不通过'
        }
        
        # 写入数据
        for row_num, row_data in enumerate(results, 2):
            ws.cell(row=row_num, column=1, value=str(row_data[0]))  # 订单ID
            ws.cell(row=row_num, column=2, value=str(row_data[1]))  # 视频ID
            ws.cell(row=row_num, column=3, value=status_map.get(row_data[2], row_data[2]))  # 状态
            ws.cell(row=row_num, column=4, value=float(row_data[3]) if row_data[3] else 0)  # 预算
            ws.cell(row=row_num, column=5, value=str(row_data[4] or ''))  # 标题
            ws.cell(row=row_num, column=6, value=str(row_data[5] or ''))  # 账号
            ws.cell(row=row_num, column=7, value=row_data[6].strftime('%Y-%m-%d %H:%M') if row_data[6] else '')  # 时间
            ws.cell(row=row_num, column=8, value=float(row_data[7]) if row_data[7] else 0)  # 消耗
            ws.cell(row=row_num, column=9, value=int(row_data[8]) if row_data[8] else 0)  # 播放量
            ws.cell(row=row_num, column=10, value=int(row_data[9]) if row_data[9] else 0)  # 点赞
            ws.cell(row=row_num, column=11, value=int(row_data[10]) if row_data[10] else 0)  # 评论
            ws.cell(row=row_num, column=12, value=int(row_data[11]) if row_data[11] else 0)  # 转发
            ws.cell(row=row_num, column=13, value=int(row_data[12]) if row_data[12] else 0)  # 关注
            ws.cell(row=row_num, column=14, value=int(row_data[13]) if row_data[13] else 0)  # 转化
            ws.cell(row=row_num, column=15, value=float(row_data[14]) if row_data[14] else 0)  # 百播放量
            ws.cell(row=row_num, column=16, value=float(row_data[15]) if row_data[15] else 0)  # 转化成本
            ws.cell(row=row_num, column=17, value=float(row_data[16]) if row_data[16] else 0)  # 百转发率
            ws.cell(row=row_num, column=18, value=float(row_data[17]) * 100 if row_data[17] else 0)  # 5秒完播率
        
        # 调整列宽
        column_widths = [20, 20, 12, 12, 40, 15, 18, 12, 12, 10, 10, 10, 10, 10, 12, 12, 12, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成文件名（使用URL编码处理中文）
        from urllib.parse import quote
        filename = f"订单数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        filename_encoded = quote(filename)
        
        # 返回文件流
        return Response(
            output.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename*=UTF-8\'\'{filename_encoded}',
                'Access-Control-Expose-Headers': 'Content-Disposition'
            }
        )
        
    except Exception as e:
        logger.error(f"导出数据失败: {e}", exc_info=True)
        return error_response(f'导出失败: {str(e)}'), 500
    finally:
        db.close()
